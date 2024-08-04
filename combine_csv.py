from openpyxl import Workbook
from openpyxl.drawing.image import Image
import json 
import cv2
import matplotlib.pyplot as plt 
from io import BytesIO
# import os 
# import re 
# import argparse
# import numpy as np 


# class CroppedDataset(): 
#     def __init__(self, img_path, json_path, save_dir, save=True):
#         self.save=save 
#         self.save_dir = save_dir
#         self.cropped_img = self.get_cropped_img(img_path, json_path)
#         self.num_cropped = len(self.cropped_img)    
#     def get_cropped_img(self, img_path, json_path): 
#         img = cv2.imread(img_path) 
#         f = open(json_path, 'r') 
#         json_data = json.load(f) 
#         cropped_data = [] 
#         for i, data in enumerate(json_data["shapes"]): 
#             try: 
#                 [[x1, y1], [x2, y2]]  = data["points"] 
#                 cropped_img = img[int(y1):int(y2), int(x1):int(x2)]
#                 f_name = os.path.basename(img_path)
#                 f_name = f"{f_name.split('.')[0]}_{str(i)}.png"
#                 f_name = os.path.join(self.save_dir, f_name)
#                 cv2.imwrite(f_name, cropped_img)
#                 cropped_data.append({
#                     "cropped_img": cropped_img, 
#                     "bbox": [int(x1), int(y1), int(x2), int(y2)], 
#                     "cropped_path": f_name  
#                 })
#             except Exception as e: 
#                 print(f"Error ocurred {e}")
#         f.close() 
#         return cropped_data

# class ImageDataset(): 
#     def __init__(self, dir, img_save): 
#         self.json_data, self.img_data = self.__getDirs(dir) 
#         self.json_data = self.__naturalSort(self.json_data) 
#         self.img_data = self.__naturalSort(self.img_data)
#         self.img_save=  img_save

#     def __getitem__(self, idx): 
#         return CroppedDataset(self.img_data[idx], self.json_data[idx], self.img_save)

#     def __getDirs(self, dir): 
#         total_json = [] 
#         total_png = [] 
#         for (root, _, files) in os.walk(dir): 
#             if files: 
#                 json_files = [os.path.join(root, i) for i in files if i.endswith(".json")]
#                 png_files = [os.path.join(root, i.split(".")[0] + ".png") for i in files if i.endswith(".json")]
#                 total_json.extend(json_files)
#                 total_png.extend(png_files)

#         assert total_json, "Files not Found"
#         return total_json, total_png 

#     def __naturalSort(self, l): 
#         convert = lambda text: int(text) if text.isdigit() else text.lower()
#         alphanum_key = lambda key: [convert(c) for c in re.split('([0-9]+)', key)]
#         return sorted(l, key=alphanum_key)

#     def genExcel(self,excel_save_dir,size = 100): 
#         data =[self.json_data[i:i+size] for i in range(0, len(self.img_data), size)]
#         for i,chunk in enumerate(data): 
#             wb = Workbook()
#             ws = wb.active
#             ws['A1'] = 'ImgPath'
#             ws['B1'] = 'CroppedPath'
#             ws['C1'] = 'Image'
#             ws['D1'] = 'Value'
#             row = 2
#             for element in chunk: 
#                 # try:
#                 idx = self.json_data.index(element)
#                 for cropped_img in self[idx].cropped_img: 
#                     if cropped_img != []: 
#                         cropped_path = cropped_img['cropped_path']
#                         cropped_img = cropped_img['cropped_img']
#                         img_height, img_width , _ = cropped_img.shape
#                         new_width = int(400)
#                         ratio = img_height / img_width
#                         new_height = int(new_width * ratio) 
#                         if img_width > 350: 
#                             cropped_img = cv2.resize(cropped_img, (new_width, new_height)) 
#                         img_bytes = cv2.imencode('.png', cropped_img)[1].tostring()
#                         img_buffer = BytesIO(img_bytes)
#                         xl_img = Image(img_buffer)
#                         ws.add_image(xl_img, f'C{row}')
#                         ws.row_dimensions[row].height = img_height
#                         ws[f'A{row}'] = element 
#                         ws[f'B{row}'] = cropped_path 
#                         ws[f'D{row}'] = " "
#                         row += 1
#                 # except Exception as e: 
#                 #     print(f"Error ocurred {e} at idx {idx} file name {element}")
#             ws.column_dimensions['A'].width = 200
#             ws.column_dimensions['B'].width = 200
#             ws.column_dimensions['C'].width = 500 / 7 
#             wb.save(f"{excel_save_dir}/Batch_{i+1}.xlsx")
#             print(f"Saving Batch_{i+1}")
#             wb.close() 


# parser = argparse.ArgumentParser()
# parser.add_argument('--image_dir', default='.', help='image dir')
# parser.add_argument('--img_save_dir', default='.', help='excel save dir')
# parser.add_argument('--excel_save_dir', default='.', help='excel save dir')

# # parser.add_argument('--save', default='data.xlsx', help='')
# # parser.add_argument('--chunk', type=int, default=100, help='Specify an integer parameter')
# opt = parser.parse_args()
# dir = opt.image_dir
# img_dir = opt.img_save_dir
# data = ImageDataset(dir, img_dir)
# data.genExcel(opt.excel_save_dir)


import pandas as pd
import os 
root_path="/root/DATA/image"
csv_path="/root/DATA/set_2_combined.csv"
final_csv = {"file_name": [], 
             "value": []
} 
final_csv_save="/root/DATA/final.csv"
test_csv = {}
df = pd.read_csv(csv_path, sep="\t", header=None) 
for i, data in enumerate(df.iloc[:]): 
    full_path = os.path.join(root_path, data[0]) 
    if os.path.exists(full_path): 
        # final_csv['file_name'].append(data[0]) 
        # final_csv['value'].append(data[1]) 
        test_csv.append
    if i == 10: 
        print(final_csv)
        final_df = pd.DataFrame.from_dict(final_csv, orient='index', columns=["file_name", "value"])
        final_df.to_csv(final_csv_save, sep="\t", header=None, index=None)
        break 
        


        
    